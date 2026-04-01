"""
workbook_parser.py

Parses Excel workbooks (.xlsx, .xls) submitted by suppliers as pricing
or response attachments.

v3.1: Now wired into response_intake_agent.py for supplier response uploads.
      Also callable from rfp.py's /rfp/upload-supplier-response endpoint.

Returns a normalised list of rows regardless of sheet structure.
"""
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# Optional dependencies — import lazily so the service starts even if libs are missing
try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
    logger.warning("openpyxl not installed — .xlsx parsing disabled")

try:
    import xlrd
    _XLRD = True
except ImportError:
    _XLRD = False
    logger.warning("xlrd not installed — .xls parsing disabled")


def parse_workbook(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Parse an Excel file and return its contents as a structured dict.

    Args:
        file_path: Absolute or relative path to the .xlsx or .xls file.
        sheet_name: If provided, parse only this sheet. Otherwise parses all sheets.

    Returns:
        {
            "sheets": {
                "Sheet1": [
                    {"col_A": value, "col_B": value, ...},  # one dict per row
                    ...
                ],
                ...
            },
            "row_count": int,
            "sheet_names": [str],
            "source": str,
        }
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(path, sheet_name)
    elif suffix == ".xls":
        return _parse_xls(path, sheet_name)
    else:
        raise ValueError(f"Unsupported workbook format: {suffix}. Expected .xlsx or .xls")


def _parse_xlsx(path: Path, sheet_name: Optional[str]) -> Dict[str, Any]:
    if not _OPENPYXL:
        raise ImportError("openpyxl is required for .xlsx parsing. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets_to_parse = [sheet_name] if sheet_name else wb.sheetnames
    result: Dict[str, List[dict]] = {}
    total_rows = 0

    for sname in sheets_to_parse:
        if sname not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in %s", sname, path.name)
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sname] = []
            continue

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = [
            {headers[i]: _clean(cell) for i, cell in enumerate(row)}
            for row in rows[1:]
            if any(cell is not None for cell in row)  # skip empty rows
        ]
        result[sname] = data
        total_rows += len(data)

    wb.close()
    return {
        "sheets": result,
        "row_count": total_rows,
        "sheet_names": list(result.keys()),
        "source": str(path),
    }


def _parse_xls(path: Path, sheet_name: Optional[str]) -> Dict[str, Any]:
    if not _XLRD:
        raise ImportError("xlrd is required for .xls parsing. Run: pip install xlrd")

    wb = xlrd.open_workbook(str(path))
    sheet_names = wb.sheet_names()
    sheets_to_parse = [sheet_name] if sheet_name else sheet_names
    result: Dict[str, List[dict]] = {}
    total_rows = 0

    for sname in sheets_to_parse:
        if sname not in sheet_names:
            logger.warning("Sheet '%s' not found in %s", sname, path.name)
            continue
        ws = wb.sheet_by_name(sname)
        if ws.nrows == 0:
            result[sname] = []
            continue

        headers = [str(ws.cell_value(0, c)).strip() or f"col_{c}" for c in range(ws.ncols)]
        data = [
            {headers[c]: _clean(ws.cell_value(r, c)) for c in range(ws.ncols)}
            for r in range(1, ws.nrows)
            if any(ws.cell_value(r, c) not in ("", None) for c in range(ws.ncols))
        ]
        result[sname] = data
        total_rows += len(data)

    return {
        "sheets": result,
        "row_count": total_rows,
        "sheet_names": list(result.keys()),
        "source": str(path),
    }


def _clean(value: Any) -> Any:
    """Normalise cell values: strip strings, convert floats that are whole numbers to int."""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def extract_pricing_rows(workbook_result: Dict[str, Any]) -> List[dict]:
    """
    Convenience function: flatten all sheets into a single list of rows
    that look like pricing data (have a numeric 'price' or 'unit_price' column).

    Used by response_intake_agent.py to extract pricing from supplier workbooks.
    """
    pricing_rows = []
    price_keys = {"price", "unit_price", "unit price", "rate", "amount", "cost"}

    for sheet_name, rows in workbook_result.get("sheets", {}).items():
        for row in rows:
            row_lower = {k.lower(): v for k, v in row.items()}
            if any(k in row_lower for k in price_keys):
                pricing_rows.append({**row, "_sheet": sheet_name})

    return pricing_rows
